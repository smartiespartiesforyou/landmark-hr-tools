import re
import unicodedata
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime

from openpyxl import load_workbook


UKG_REQUIRED = {"Employee Id", "Last Name", "First Name", "Employee Status", "Date Hired", "Date Re-Hired", "DEPT"}
TERMINATED_STATUSES = {"terminated", "deceased", "resigned", "retired"}
COMPLETED_MARKS = {"✔", "✓"}
KNOWN_DUPLICATE_NAMES = {
    "bridget horn",
    "tequira millican",
    "tia octave",
    "tina holmes",
}


def normalize_name(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = value.encode("ascii", "ignore").decode("ascii").casefold().strip()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _find_header_row(rows):
    for index, row in enumerate(rows):
        values = {str(value).strip() for value in row if value is not None}
        if {"Last Name", "First Name", "Date Hired"}.issubset(values):
            return index
    raise ValueError("Could not find the employee header row in the UKG file.")


def read_ukg_employees(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_row = _find_header_row(rows)
    headers = [str(value).strip() if value is not None else "" for value in rows[header_row]]
    missing = UKG_REQUIRED.difference(headers)
    if missing:
        workbook.close()
        raise ValueError("UKG file is missing: " + ", ".join(sorted(missing)))

    column = {name: index for index, name in enumerate(headers)}
    employees = []
    for row in rows[header_row + 1:]:
        employee_id = row[column["Employee Id"]]
        if employee_id in (None, ""):
            continue
        status = str(row[column["Employee Status"]] or "").strip()
        if status.casefold() in TERMINATED_STATUSES:
            continue
        first = str(row[column["First Name"]] or "").strip()
        last = str(row[column["Last Name"]] or "").strip()
        employees.append({
            "employee_id": str(employee_id).strip(),
            "first": first,
            "last": last,
            "display_name": f"{first} {last}".strip().title(),
            "name_keys": {normalize_name(f"{first} {last}"), normalize_name(f"{last} {first}")},
            "department": str(row[column["DEPT"]] or "").strip(),
            "job": str(row[column["JOB"]] or "").strip() if "JOB" in column else "",
            "effective_hire": _as_date(row[column["Date Re-Hired"]]) or _as_date(row[column["Date Hired"]]),
        })
    workbook.close()
    return employees


def read_training_matrix(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Training matrix" not in workbook.sheetnames:
        workbook.close()
        raise ValueError("TalentLMS file is missing the Training matrix sheet.")
    sheet = workbook["Training matrix"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers or str(headers[0] or "").strip() != "User":
        workbook.close()
        raise ValueError("TalentLMS Training Matrix must begin with a User column.")
    courses = [
        {"index": index, "name": str(value).strip()}
        for index, value in enumerate(headers[1:], start=1)
        if value not in (None, "")
    ]
    matrix_rows = []
    for row in rows:
        if not row or row[0] in (None, ""):
            continue
        matrix_rows.append({"user": str(row[0]).strip(), "key": normalize_name(row[0]), "values": tuple(row)})
    workbook.close()
    return courses, matrix_rows


def match_employees(employees, matrix_rows):
    rows_by_key = defaultdict(list)
    for row in matrix_rows:
        rows_by_key[row["key"]].append(row)
    unique, duplicates, missing = [], [], []
    for employee in employees:
        hits_by_identity = {}
        for key in employee["name_keys"]:
            for row in rows_by_key.get(key, []):
                hits_by_identity[id(row)] = row
        hits = list(hits_by_identity.values())
        if len(hits) == 1:
            unique.append((employee, hits[0]))
        elif len(hits) > 1:
            # TalentLMS will not remove four known duplicate accounts. HR marks the
            # correct account by capitalizing the entire TalentLMS user name.
            employee_key = normalize_name(f"{employee['first']} {employee['last']}")
            if employee_key in KNOWN_DUPLICATE_NAMES:
                # Prefer the account HR marked in ALL CAPS when TalentLMS preserves
                # that capitalization in the Training Matrix export.
                capitalized_hits = [row for row in hits if row["user"].isupper()]
                if len(capitalized_hits) == 1:
                    unique.append((employee, capitalized_hits[0]))
                    continue

                # Some TalentLMS exports title-case duplicate names (for example,
                # both Bridget Horn rows export as "Horn Bridget"). For these four
                # known duplicate employees only, fall back to the account with the
                # most recorded training activity. Require a single clear winner.
                def activity_count(row):
                    return sum(value not in (None, "") for value in row["values"][1:])

                ranked_hits = sorted(hits, key=activity_count, reverse=True)
                if len(ranked_hits) >= 2 and activity_count(ranked_hits[0]) > activity_count(ranked_hits[1]):
                    unique.append((employee, ranked_hits[0]))
                    continue
            duplicates.append({"employee": employee, "matrix_names": [row["user"] for row in hits]})
        else:
            missing.append(employee)
    return unique, duplicates, missing


def course_month_end(course_name):
    match = re.match(r"^\s*(0?[1-9]|1[0-2])\s+[A-Za-z]+\b.*\b(20\d{2})\s*$", str(course_name or ""))
    if not match:
        return None
    month, year = int(match.group(1)), int(match.group(2))
    return date(year, month, monthrange(year, month)[1])


def is_completed(matrix_row, course_index):
    values = matrix_row["values"]
    return course_index < len(values) and str(values[course_index] or "").strip() in COMPLETED_MARKS


def filter_eligible(employees, scope, departments, jobs, eligibility_end=None):
    selected = []
    department_set = {str(value).strip().upper() for value in (departments or [])}
    job_set = {str(value).strip().upper() for value in (jobs or [])}
    for employee in employees:
        department = employee["department"].strip().upper()
        job = employee["job"].strip().upper()
        if scope == "departments":
            # CNA is offered as a department-style print choice even though UKG
            # stores CNA as the job under the NSG department.
            matches_cna = "CNA" in department_set and department == "NSG" and job == "CNA"
            matches_department = department in department_set
            if not (matches_cna or matches_department):
                continue
        if scope == "jobs" and job not in job_set:
            continue
        if eligibility_end and employee["effective_hire"] and employee["effective_hire"] > eligibility_end:
            continue
        selected.append(employee)
    return selected
