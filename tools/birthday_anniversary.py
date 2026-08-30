import re
from datetime import date, datetime

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


INACTIVE_STATUSES = {"terminated", "deceased", "resigned", "retired", "inactive"}


def _clean_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _column(columns, aliases, label, required=True):
    for alias in aliases:
        if alias in columns:
            return columns[alias]
    if required:
        raise ValueError(f"UKG report is missing the {label} column.")
    return None


def _find_header_row(rows):
    for index, row in enumerate(rows[:30]):
        headers = {_clean_header(value) for value in row}
        if {"firstname", "lastname", "latestdate"}.issubset(headers):
            return index
    raise ValueError("Could not find the employee column headings in the UKG report.")


def read_birthdays_and_anniversaries(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
        header_row = _find_header_row(rows)
        columns = {
            _clean_header(value): index
            for index, value in enumerate(rows[header_row])
            if value not in (None, "")
        }
        first_col = _column(columns, {"firstname"}, "First Name")
        last_col = _column(columns, {"lastname"}, "Last Name")
        birthday_col = _column(columns, {"datebirthday", "dateofbirth", "birthdate", "dob"}, "Date Birthday")
        latest_col = _column(columns, {"latestdate", "latesthireorrehiredate"}, "Latest Date")
        status_col = _column(columns, {"employeestatus", "status"}, "Employee Status", required=False)

        employees = []
        missing_birthdays = []
        for row in rows[header_row + 1:]:
            first = str(row[first_col] or "").strip()
            last = str(row[last_col] or "").strip()
            if not first and not last:
                continue
            status = str(row[status_col] or "").strip() if status_col is not None else ""
            if status.casefold() in INACTIVE_STATUSES:
                continue
            birthday = _as_date(row[birthday_col])
            anniversary = _as_date(row[latest_col])
            name = f"{first.title()} {last.title()}".strip()
            if birthday is None:
                missing_birthdays.append(name)
                continue
            if anniversary is None:
                continue
            employees.append({"name": name, "birthday": birthday, "anniversary": anniversary})
        return employees, missing_birthdays
    finally:
        workbook.close()


def _table(data, widths, compact=False):
    table = Table(data, repeatRows=1, colWidths=widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#174F79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5 if compact else 9.5),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#777777")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5F7")]),
        ("TOPPADDING", (0, 0), (-1, -1), 0.5 if compact else 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0.5 if compact else 6),
    ]))
    return table


def create_birthday_anniversary_pdf(employees, output_path, month=None):
    document = SimpleDocTemplate(
        output_path,
        pagesize=landscape(letter) if month is not None else letter,
        rightMargin=0.55 * inch,
        leftMargin=0.55 * inch,
        topMargin=0.48 * inch,
        bottomMargin=0.48 * inch,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "CelebrationTitle",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=16,
        leading=19,
        spaceAfter=12,
    )
    section = ParagraphStyle(
        "CelebrationSection",
        parent=styles["Heading2"],
        fontSize=12,
        leading=14,
        spaceBefore=4,
        spaceAfter=6,
    )
    story = []

    if month is not None:
        month_name = date(2000, month, 1).strftime("%B")
        story.append(Paragraph(f"{month_name} Birthdays and Work Anniversaries", title))

        birthdays = sorted(
            (item for item in employees if item["birthday"].month == month),
            key=lambda item: (item["birthday"].day, item["name"].casefold()),
        )
        anniversaries = sorted(
            (item for item in employees if item["anniversary"].month == month),
            key=lambda item: (item["anniversary"].day, item["name"].casefold()),
        )

        birthday_data = [["Employee", "Birthday"]]
        birthday_data.extend([[item["name"], item["birthday"].strftime("%m/%d")] for item in birthdays])
        if not birthdays:
            birthday_data.append(["No birthdays this month", ""])
        anniversary_data = [["Employee", "Anniversary"]]
        anniversary_data.extend([[item["name"], item["anniversary"].strftime("%m/%d")] for item in anniversaries])
        if not anniversaries:
            anniversary_data.append(["No work anniversaries this month", ""])
        birthday_table = _table(birthday_data, [3.8 * inch, 0.8 * inch], compact=True)
        anniversary_table = _table(anniversary_data, [3.8 * inch, 0.8 * inch], compact=True)
        board = Table(
            [[
                [Paragraph("Birthdays", section), birthday_table],
                [Paragraph("Work Anniversaries", section), anniversary_table],
            ]],
            colWidths=[4.75 * inch, 4.75 * inch],
            hAlign="CENTER",
        )
        board.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(board)
    else:
        story.append(Paragraph("Employee Birthday and Work Anniversary Dates", title))
        story.append(Paragraph("Active employees - birthday years are intentionally omitted", styles["Normal"]))
        story.append(Spacer(1, 10))
        data = [["Employee", "Birthday", "Work Anniversary"]]
        for item in sorted(employees, key=lambda employee: employee["name"].casefold()):
            data.append([
                item["name"],
                item["birthday"].strftime("%m/%d"),
                item["anniversary"].strftime("%m/%d"),
            ])
        story.append(_table(data, [4.5 * inch, 1.2 * inch, 1.2 * inch]))

    document.build(story)
