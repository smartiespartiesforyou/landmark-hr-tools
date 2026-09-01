import re
from calendar import monthrange
from datetime import date, datetime, timedelta

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


TERMINATED_STATUSES = {"terminated", "deceased", "resigned", "retired", "inactive"}

# Department codes whose evaluations should be combined into James's packet.
JAMES_DEPARTMENT_CODES = {"ADM", "SSV", "ACT", "MTN", "MRD", "NCL", "ASL", "ADL", "ADR", "ASR", "DON"}


def _clean_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _column(columns, aliases, label, required=True):
    for alias in aliases:
        if alias in columns:
            return columns[alias]
    if required:
        raise ValueError(f"UKG report is missing the {label} column.")
    return None


def _find_header_row(rows):
    for index, row in enumerate(rows[:30]):
        headers = {_clean_header(value) for value in row}
        if {"firstname", "lastname"}.issubset(headers) and (
            "latestdate" in headers or "datehired" in headers
        ):
            return index
    raise ValueError("Could not find the employee column headings in the UKG report.")


def read_evaluation_employees(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header_row = _find_header_row(rows)
        columns = {
            _clean_header(value): index
            for index, value in enumerate(rows[header_row])
            if value not in (None, "")
        }

        first_col = _column(columns, {"firstname"}, "First Name")
        last_col = _column(columns, {"lastname"}, "Last Name")
        department_col = _column(
            columns,
            {"dept", "department", "departmentname", "homedepartment", "location4"},
            "Department (DEPT)",
        )
        job_col = _column(
            columns,
            {"job", "primaryjob", "jobtitle", "position", "positiontitle", "location5"},
            "Primary Job/Position (JOB)",
        )
        latest_col = _column(columns, {"latestdate", "latesthireorrehiredate"}, "Latest Date", required=False)
        hire_col = _column(columns, {"datehired", "hiredate"}, "Date Hired", required=False)
        rehire_col = _column(columns, {"daterehired", "rehiredate"}, "Date Re-Hired", required=False)
        status_col = _column(columns, {"employeestatus", "status"}, "Employee Status", required=False)

        if latest_col is None and hire_col is None:
            raise ValueError("UKG report is missing Latest Date and Date Hired.")

        employees = []
        for row in rows[header_row + 1:]:
            first = str(row[first_col] or "").strip()
            last = str(row[last_col] or "").strip()
            if not first and not last:
                continue
            status = str(row[status_col] or "").strip() if status_col is not None else ""
            if status.casefold() in TERMINATED_STATUSES:
                continue
            latest = _as_date(row[latest_col]) if latest_col is not None else None
            if latest is None and rehire_col is not None:
                latest = _as_date(row[rehire_col])
            if latest is None and hire_col is not None:
                latest = _as_date(row[hire_col])
            if latest is None:
                continue
            department = str(row[department_col] or "").strip() or "Department Not Listed"
            job = str(row[job_col] or "").strip() or "Not Listed"

            # Keep CNAs on their own evaluation printout instead of combining
            # them with nurses under the NSG department.
            department_code = department.upper()
            job_code = job.upper()
            if department_code == "NSG" and job_code == "CNA":
                department = "CNAs"
            elif department_code == "NSG":
                department = "Nursing"
            elif department_code in JAMES_DEPARTMENT_CODES:
                department = "James"

            employees.append({
                "name": f"{first.title()} {last.title()}".strip(),
                "department": department,
                "job": job,
                "latest_date": latest,
            })
        return employees
    finally:
        workbook.close()


def _annual_due_date(hire_date, year):
    day = min(hire_date.day, monthrange(year, hire_date.month)[1])
    return date(year, hire_date.month, day)


def evaluations_for_month(employees, year, month):
    results = []
    for employee in employees:
        hire_date = employee["latest_date"]
        ninety_day = hire_date + timedelta(days=90)
        if ninety_day.year == year and ninety_day.month == month:
            results.append({**employee, "type": "90-Day", "due_date": ninety_day})

        annual = _annual_due_date(hire_date, year)
        if annual.month == month and annual >= hire_date + timedelta(days=365):
            results.append({**employee, "type": "Annual", "due_date": annual})

    return sorted(
        results,
        key=lambda item: (
            item["department"].casefold(),
            item["due_date"],
            item["name"].casefold(),
            item["type"],
        ),
    )


def create_evaluation_pdf(rows, year, month, output_path):
    report_month = date(year, month, 1).strftime("%B %Y")
    document = SimpleDocTemplate(
        output_path,
        pagesize=landscape(letter),
        rightMargin=0.42 * inch,
        leftMargin=0.42 * inch,
        topMargin=0.38 * inch,
        bottomMargin=0.38 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "EvaluationTitle",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=16,
        leading=19,
        spaceAfter=6,
    )
    department_style = ParagraphStyle(
        "Department",
        parent=styles["Heading2"],
        fontSize=12,
        leading=14,
        spaceAfter=4,
    )
    small = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=8.5, leading=10)

    departments = {}
    for row in rows:
        departments.setdefault(row["department"], []).append(row)

    story = []
    for department_index, department in enumerate(sorted(departments, key=str.casefold)):
        if department_index:
            story.append(PageBreak())
        department_rows = departments[department]
        story.append(Paragraph("Employee Evaluations Due", title_style))
        story.append(Paragraph(f"<b>{report_month}</b>", styles["Normal"]))
        story.append(Spacer(1, 5))
        story.append(Paragraph(f"Department: <b>{department}</b>", department_style))
        story.append(Paragraph("Department Manager: ____________________________________", styles["Normal"]))
        story.append(Spacer(1, 9))

        data = [[
            "Employee",
            "Position",
            "Hire/Rehire Date",
            "Evaluation Type",
            "Due Date",
            "Returned",
        ]]
        for row in department_rows:
            data.append([
                Paragraph(row["name"], small),
                Paragraph(row["job"], small),
                row["latest_date"].strftime("%m/%d/%Y"),
                row["type"],
                row["due_date"].strftime("%m/%d/%Y"),
                "[   ]",
            ])

        table = Table(
            data,
            repeatRows=1,
            colWidths=[2.0 * inch, 2.0 * inch, 1.35 * inch, 1.35 * inch, 1.15 * inch, 0.85 * inch],
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#174F79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ALIGN", (2, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#777777")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5F7")]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))
        story.append(Paragraph(
            f"Total evaluations due: <b>{len(department_rows)}</b>",
            styles["Normal"],
        ))

    document.build(story)
