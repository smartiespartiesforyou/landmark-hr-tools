from collections import defaultdict
from datetime import date, datetime
import re

from openpyxl import load_workbook

ATTENDANCE_REVIEW_VERSION = "2026-08-21-DOUBLE-LUNCH-V4"
MINIMUM_LUNCH_SHIFT_HOURS = 7
DOUBLE_SHIFT_HOURS = 12
MINIMUM_INFERRED_LUNCH_MINUTES = 20
MAXIMUM_INFERRED_LUNCH_MINUTES = 120
LUNCH_REQUIRED_EMPLOYEES = {
    "MICHELLI, COURTNEY",
    "TRAVELER, AYANA",
}
ADMIN_EMPLOYEES = {
    "BOWERS, BRIDGETT", "BREWSTER, HEATHER", "BROWN, JENIQUE", "BROWN, RAMONA",
    "CLARK, QUIAONTA", "DORSEY, MARVELETT", "DUFOUR, DOLORES", "DUGAS, DANIELLE",
    "EASTERLING, ETHAN", "HORN, BRIDGET", "JOHNSON, THERESA", "JOSEPH, JEFFREY",
    "LAMBERT, MARY", "MAJLESEIN, FAREBA", "MANCUSO, HEATHER", "MICHELLI, COURTNEY",
    "OCTAVE, TIA", "PERKINS, CHARISMA", "PRIEST, ANTOINETTE", "ROBINSON, SHONTA",
    "SPANN, LIONEL", "STROTHER, ELIZABETH", "TRAVELER, AYANA", "WALKER, SHARRON",
    "WALTON, LARA", "WILLIAMS, BRANDI",
}


def _clean(value):
    return "" if value is None else str(value).strip()


def _as_number(value):
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(_clean(value))
    except ValueError:
        return 0.0


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for pattern in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(_clean(value), pattern).date()
        except ValueError:
            pass
    raise ValueError(f"Could not read report date: {value}")


def _as_datetime(value):
    if isinstance(value, datetime):
        return value
    text = re.sub(r"\s+", " ", _clean(value))
    if not text or text == "-":
        return None
    text = re.sub(r"(\d)([ap])$", lambda match: match.group(1) + match.group(2).upper() + "M", text, flags=re.IGNORECASE)
    for pattern in ("%m/%d/%Y %I:%M%p", "%m/%d/%Y %I:%M %p", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text.upper(), pattern)
        except ValueError:
            pass
    return None


def _header_map(row):
    return {re.sub(r"\s+", " ", _clean(value)): i for i, value in enumerate(row) if _clean(value)}


def _find_header(sheet, required_aliases, report_name):
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        available = _header_map(row)
        columns = {}
        for field, aliases in required_aliases.items():
            match = next((available[alias] for alias in aliases if alias in available), None)
            if match is None:
                break
            columns[field] = match
        if len(columns) == len(required_aliases):
            return row_number, columns
    raise ValueError(f"The {report_name} headings were not recognized. Download the UKG report without renaming its columns.")


def _name_key(last, first):
    return f"{re.sub(r'\s+', ' ', _clean(last)).upper()}, {re.sub(r'\s+', ' ', _clean(first)).upper()}"


def _display_name(last, first):
    return f"{_clean(first).title()} {_clean(last).title()}".strip()


def _department(name_key, location_4, location_5):
    location_4 = _clean(location_4).upper()
    location_5 = _clean(location_5).upper()
    if name_key in ADMIN_EMPLOYEES:
        return "Administrator"
    if location_4 == "NSG" and location_5 == "CNA":
        return "CNAs"
    return {"DTY": "Dietary", "HSK": "Housekeeping", "LDY": "Laundry", "NSG": "Nursing"}.get(location_4, location_4 or "Department Not Listed")


def _format_clock(value):
    if value is None:
        return "-"
    return f"{value.hour % 12 or 12}:{value.minute:02d} {'AM' if value.hour < 12 else 'PM'}"


def _is_overnight_cna(location_4, location_5, first_in, final_out):
    location_4 = _clean(location_4).upper()
    location_5 = _clean(location_5).upper()
    is_cna_or_nat = (location_4 == "NSG" and location_5 == "CNA") or location_4 == "NAT"
    if not is_cna_or_nat:
        return False
    duration = (final_out - first_in).total_seconds() / 3600
    return first_in.hour >= 21 and final_out.hour <= 7 and 6 <= duration <= 10


def _inferred_lunch_entries(rows):
    ordered = sorted(
        (r for r in rows if r["clock_in"] is not None and r["clock_out"] is not None),
        key=lambda r: r["clock_in"],
    )
    entries = []
    for earlier, later in zip(ordered, ordered[1:]):
        gap_minutes = (later["clock_in"] - earlier["clock_out"]).total_seconds() / 60
        exceptions = f"{earlier['exceptions']} {later['exceptions']}".upper()
        is_ukg_break_exception = "MB OVER 8" in exceptions or "LONG INTRV" in exceptions
        if is_ukg_break_exception and MINIMUM_INFERRED_LUNCH_MINUTES <= gap_minutes <= MAXIMUM_INFERRED_LUNCH_MINUTES:
            entries.append(gap_minutes / 60)
    return entries


def _expected_breaks(rows):
    completed = [r for r in rows if r["clock_in"] is not None and r["clock_out"] is not None]
    if len(completed) > 1:
        return max(1, sum(2 if r["worked"] >= DOUBLE_SHIFT_HOURS else 1 for r in completed if r["worked"] >= MINIMUM_LUNCH_SHIFT_HOURS))
    total = sum(r["worked"] for r in completed)
    return 2 if total >= DOUBLE_SHIFT_HOURS else 1


def read_worked_time_report(path):
    sheet = load_workbook(path, data_only=True, read_only=True).active
    header_row, columns = _find_header(sheet, {
        "Employee Id": ("Employee Id",), "First Name": ("First Name",), "Last Name": ("Last Name",),
        "Date": ("Date",), "In": ("In Date Time (Raw)",), "Out": ("Out Date Time (Raw)",),
        "Worked": ("Total Work Hours",), "Location 4": ("Location(4)",), "Location 5": ("Location(5)",),
        "Exceptions": ("Exceptions",),
    }, "Calculated Hours By Work Day report")
    grouped = defaultdict(list)
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        employee_id = _clean(row[columns["Employee Id"]])
        first, last, raw_date = _clean(row[columns["First Name"]]), _clean(row[columns["Last Name"]]), row[columns["Date"]]
        if not employee_id or not first or not last or raw_date in (None, ""):
            continue
        clock_in, clock_out = _as_datetime(row[columns["In"]]), _as_datetime(row[columns["Out"]])
        if clock_in is None and clock_out is None:
            continue
        grouped[(employee_id, _as_date(raw_date))].append({
            "first": first, "last": last, "clock_in": clock_in, "clock_out": clock_out,
            "worked": _as_number(row[columns["Worked"]]),
            "location_4": _clean(row[columns["Location 4"]]), "location_5": _clean(row[columns["Location 5"]]),
            "exceptions": _clean(row[columns["Exceptions"]]),
        })
    if not grouped:
        raise ValueError("No worked employee rows were found in the Calculated Hours report.")
    return grouped


def read_lunch_report(path):
    sheet = load_workbook(path, data_only=True, read_only=True).active
    header_row, columns = _find_header(sheet, {
        "First Name": ("First Name",), "Last Name": ("Last Name",), "Date": ("Date",),
        "Unpaid Hours": ("Unpaid Hours", "Lunch/Break: Unpaid Hours"),
    }, "Lunch & Breaks Report")
    lunches, dates = defaultdict(list), set()
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        first, last, raw_date = _clean(row[columns["First Name"]]), _clean(row[columns["Last Name"]]), row[columns["Date"]]
        if not first or not last or raw_date in (None, ""):
            continue
        work_date = _as_date(raw_date)
        dates.add(work_date)
        hours = _as_number(row[columns["Unpaid Hours"]])
        if hours > 0:
            lunches[(_name_key(last, first), work_date)].append(hours)
    return lunches, dates


def build_daily_review(worked_time_path, lunch_path):
    worked_time = read_worked_time_report(worked_time_path)
    lunches, lunch_dates = read_lunch_report(lunch_path)
    worked_dates = {work_date for _, work_date in worked_time}
    if len(worked_dates) != 1:
        raise ValueError("The Calculated Hours report must cover exactly one day.")
    report_date = next(iter(worked_dates))
    if lunch_dates and lunch_dates != {report_date}:
        raise ValueError("The two reports are not for the same date.")

    results = []
    for (_, work_date), rows in worked_time.items():
        first, last = rows[0]["first"], rows[0]["last"]
        name_key = _name_key(last, first)
        location_4 = next((r["location_4"] for r in rows if r["location_4"]), "")
        location_5 = next((r["location_5"] for r in rows if r["location_5"]), "")
        clock_ins = [r["clock_in"] for r in rows if r["clock_in"] is not None]
        clock_outs = [r["clock_out"] for r in rows if r["clock_out"] is not None]
        has_incomplete_segment = any(r["clock_in"] is None or r["clock_out"] is None for r in rows)
        if has_incomplete_segment or not clock_ins or not clock_outs:
            results.append({"department": "Incomplete Punches / Still Working", "employee": _display_name(last, first), "date": work_date, "clock_in": "-", "clock_out": "-", "lunch": "Review", "lunch_minutes": "-"})
            continue

        first_in, final_out = min(clock_ins), max(clock_outs)
        worked_hours = sum(r["worked"] for r in rows)
        lunch_entries = lunches.get((name_key, work_date), [])
        if not lunch_entries:
            lunch_entries = _inferred_lunch_entries(rows)
        lunch_minutes = int(round(sum(lunch_entries) * 60))
        expected_breaks = _expected_breaks(rows)
        nurse_exempt = (
            location_5.upper() in {"LPN", "RGN"}
            and name_key not in LUNCH_REQUIRED_EMPLOYEES
            and name_key not in ADMIN_EMPLOYEES
        )
        lunch_exempt = nurse_exempt or _is_overnight_cna(location_4, location_5, first_in, final_out)
        lunch_required = worked_hours >= MINIMUM_LUNCH_SHIFT_HOURS and not lunch_exempt
        if lunch_entries:
            if expected_breaks == 2 and len(lunch_entries) == 1 and lunch_minutes < 60:
                lunch_status = "1 of 2"
            else:
                lunch_status = "Yes"
        elif worked_hours < MINIMUM_LUNCH_SHIFT_HOURS:
            lunch_status = "Short Shift"
        else:
            lunch_status = "No" if lunch_required else "N/A"
        results.append({
            "department": _department(name_key, location_4, location_5), "employee": _display_name(last, first),
            "date": work_date, "clock_in": _format_clock(first_in), "clock_out": _format_clock(final_out),
            "lunch": lunch_status, "lunch_minutes": lunch_minutes if lunch_minutes else "-",
        })
    results.sort(key=lambda r: (r["department"] == "Incomplete Punches / Still Working", r["department"], r["employee"]))
    return report_date, results, []
